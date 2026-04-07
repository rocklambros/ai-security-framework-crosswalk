"""Excel writer: 5-sheet workbook (README + 4 data sheets).

Framework-agnostic port of ``AIUC_2_OWASP_Agentic_Top_10/aiuc/output.py``.
All labels come from ``pair_config`` rather than hard-coded AIUC/OWASP text.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from mapping_engine.config import PairConfig
from mapping_engine.engine.mapper import MappingResult

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
DIRECT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RELATED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PRIMARY_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PRIMARY_FONT = Font(color="FFFFFF")
SECONDARY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _label(framework: str) -> str:
    return {"aiuc_1": "AIUC-1", "owasp_agentic": "OWASP Agentic"}.get(
        framework, framework.replace("_", " ").title()
    )


def _style_header(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER


def _auto_width(ws, n_cols: int, max_rows: int) -> None:
    for col in range(1, n_cols + 1):
        mx = 0
        for r in range(1, min(max_rows + 1, 100)):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 2, 50)


def _tier_fill(ws, col: int, start: int, end: int) -> None:
    for r in range(start, end + 1):
        c = ws.cell(row=r, column=col)
        v = str(c.value or "")
        if v == "Direct":
            c.fill = DIRECT_FILL
        elif v == "Related":
            c.fill = RELATED_FILL
        c.border = BORDER


def _relevance_fill(ws, col: int, start: int, end: int) -> None:
    for r in range(start, end + 1):
        c = ws.cell(row=r, column=col)
        v = str(c.value or "")
        if v == "Primary":
            c.fill = PRIMARY_FILL
            c.font = PRIMARY_FONT
        elif v == "Secondary":
            c.fill = SECONDARY_FILL
        c.border = BORDER


def _write_readme(wb: Workbook, pair_config: PairConfig, result: MappingResult) -> None:
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.cell(row=1, column=1, value=f"{_label(pair_config.source_framework)} ↔ {_label(pair_config.target_framework)} Crosswalk").font = Font(bold=True, size=14, color="1F4E79")
    ws.cell(row=3, column=1, value="Generated at:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=result.metadata["generated_at"])
    ws.cell(row=4, column=1, value="Methodology:").font = Font(bold=True)
    ws.cell(row=4, column=2, value=result.metadata.get("methodology", ""))
    ws.cell(row=5, column=1, value="Total mappings:").font = Font(bold=True)
    ws.cell(row=5, column=2, value=len(result.mappings))

    ws.cell(row=7, column=1, value="Tier colors:").font = Font(bold=True)
    ws.cell(row=8, column=1, value="Direct").fill = DIRECT_FILL
    ws.cell(row=8, column=2, value="Strong, well-supported connection (composite >= direct threshold)")
    ws.cell(row=9, column=1, value="Related").fill = RELATED_FILL
    ws.cell(row=9, column=2, value="Meaningful but less direct connection")

    ws.cell(row=11, column=1, value="Rationale codes:").font = Font(bold=True)
    rows = [
        ("PREV", "Prevent — blocks the core attack mechanism"),
        ("SCOPE", "Constrain scope — limits blast radius"),
        ("GATE", "Human gate — enforces human approval"),
        ("DETECT", "Detect and trace"),
        ("VALID", "Validate and test"),
        ("GOVERN", "Policy and governance"),
        ("ISOLATE", "Isolate and contain"),
        ("DISCLOSE", "Disclose and calibrate"),
    ]
    for i, (code, desc) in enumerate(rows, start=12):
        ws.cell(row=i, column=1, value=code).font = Font(bold=True)
        ws.cell(row=i, column=2, value=desc)


def _write_source_to_target(
    wb: Workbook,
    result: MappingResult,
    G: nx.DiGraph,
    pair_config: PairConfig,
    sheet_name: str,
    entry_types: list[str] | None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    headers = [
        f"{_label(pair_config.source_framework)} ID",
        "Title", "Domain", "Function Class",
        f"{_label(pair_config.target_framework)} ID", "Target Title",
        "Rationale", "Relevance", "Score", "Tier",
        "Bridge", "Semantic", "Keyword", "Function Match",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    row = 2
    for m in result.mappings:
        s_node = G.nodes[m["source_node_id"]]
        t_node = G.nodes[m["target_node_id"]]
        if entry_types is not None and s_node.get("entry_type") not in entry_types:
            continue
        ws.cell(row=row, column=1, value=s_node.get("local_id") or m["source_node_id"])
        ws.cell(row=row, column=2, value=s_node.get("name") or "")
        ws.cell(row=row, column=3, value=s_node.get("domain") or "")
        ws.cell(row=row, column=4, value=s_node.get("function_class") or m["rationale_code"])
        ws.cell(row=row, column=5, value=t_node.get("local_id") or m["target_node_id"])
        ws.cell(row=row, column=6, value=t_node.get("name") or "")
        ws.cell(row=row, column=7, value=m["rationale_code"])
        ws.cell(row=row, column=8, value=m["relevance"])
        ws.cell(row=row, column=9, value=round(m["score"], 4))
        ws.cell(row=row, column=10, value=m["tier"])
        sig = m["signals"]
        ws.cell(row=row, column=11, value=round(sig.get("bridge", 0.0), 4))
        ws.cell(row=row, column=12, value=round(sig.get("semantic", 0.0), 4))
        ws.cell(row=row, column=13, value=round(sig.get("keyword", 0.0), 4))
        ws.cell(row=row, column=14, value=round(sig.get("function_match", 0.0), 4))
        row += 1
    _tier_fill(ws, 10, 2, row - 1)
    _relevance_fill(ws, 8, 2, row - 1)
    _auto_width(ws, len(headers), row - 1)


def _write_target_to_source(
    wb: Workbook,
    result: MappingResult,
    G: nx.DiGraph,
    pair_config: PairConfig,
    sheet_name: str,
    entry_types: list[str] | None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    headers = [
        f"{_label(pair_config.target_framework)} ID", "Target Title",
        "Function Coverage",
        f"{_label(pair_config.source_framework)} ID", "Source Title",
        "Function Class", "Rationale", "Relevance", "Score", "Tier",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    by_tgt: dict[str, list[dict[str, Any]]] = {}
    for m in result.mappings:
        s_node = G.nodes[m["source_node_id"]]
        if entry_types is not None and s_node.get("entry_type") not in entry_types:
            continue
        by_tgt.setdefault(m["target_node_id"], []).append(m)

    row = 2
    for t in result.target_nodes:
        t_node = G.nodes[t]
        ms = sorted(by_tgt.get(t, []), key=lambda x: -x["score"])
        cov: dict[str, int] = {}
        for m in ms:
            cov[m["rationale_code"]] = cov.get(m["rationale_code"], 0) + 1
        cov_str = ", ".join(f"{k}:{v}" for k, v in sorted(cov.items(), key=lambda x: -x[1])) or "—"
        if not ms:
            ws.cell(row=row, column=1, value=t_node.get("local_id") or t)
            ws.cell(row=row, column=2, value=t_node.get("name") or "")
            ws.cell(row=row, column=3, value="—")
            row += 1
            continue
        for m in ms:
            s_node = G.nodes[m["source_node_id"]]
            ws.cell(row=row, column=1, value=t_node.get("local_id") or t)
            ws.cell(row=row, column=2, value=t_node.get("name") or "")
            ws.cell(row=row, column=3, value=cov_str)
            ws.cell(row=row, column=4, value=s_node.get("local_id") or m["source_node_id"])
            ws.cell(row=row, column=5, value=s_node.get("name") or "")
            ws.cell(row=row, column=6, value=s_node.get("function_class") or m["rationale_code"])
            ws.cell(row=row, column=7, value=m["rationale_code"])
            ws.cell(row=row, column=8, value=m["relevance"])
            ws.cell(row=row, column=9, value=round(m["score"], 4))
            ws.cell(row=row, column=10, value=m["tier"])
            row += 1
    _tier_fill(ws, 10, 2, row - 1)
    _relevance_fill(ws, 8, 2, row - 1)
    _auto_width(ws, len(headers), row - 1)


def write_excel(
    result: MappingResult,
    G: nx.DiGraph,
    pair_config: PairConfig,
    output_path: str | Path,
) -> Path:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _write_readme(wb, pair_config, result)
    _write_source_to_target(
        wb, result, G, pair_config,
        f"{_label(pair_config.source_framework)}→{_label(pair_config.target_framework)} (Control)",
        entry_types=["control"],
    )
    _write_target_to_source(
        wb, result, G, pair_config,
        f"{_label(pair_config.target_framework)}→{_label(pair_config.source_framework)} (Control)",
        entry_types=["control"],
    )
    _write_source_to_target(
        wb, result, G, pair_config,
        f"{_label(pair_config.source_framework)}→{_label(pair_config.target_framework)} (Activity)",
        entry_types=["activity"],
    )
    _write_target_to_source(
        wb, result, G, pair_config,
        f"{_label(pair_config.target_framework)}→{_label(pair_config.source_framework)} (Activity)",
        entry_types=["activity"],
    )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
