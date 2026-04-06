#!/usr/bin/env python3
"""
Convert CSA AI Controls Matrix (AICM) v1.0.3 spreadsheet to JSON and markdown.

Source: AICMv1.0.3 xlsx bundle from CSA.
Extracts controls from the AICM sheet (30 cols), cross-framework mappings
from Scope Applicability (Mappings) sheet, and implementation/auditing
guidelines from their respective sheets.

Outputs:
  - csa_aicm.json: Structured JSON of all controls with mappings
  - csa_aicm.md: Human-readable markdown reference
"""

import json
import openpyxl
from pathlib import Path
from collections import OrderedDict

AICM_DIR = Path("data/frameworks/csa-aicm")


def find_main_xlsx():
    """Find the main AICM xlsx file (not the AI-CAIQ)."""
    xlsx_files = sorted(AICM_DIR.glob("*.xlsx"))
    for f in xlsx_files:
        if "aicm" in f.name.lower() and "caiq" not in f.name.lower():
            return f
    # Fallback: largest xlsx
    if xlsx_files:
        return max(xlsx_files, key=lambda f: f.stat().st_size)
    return None


def find_caiq_xlsx():
    """Find the AI-CAIQ xlsx file."""
    for f in sorted(AICM_DIR.glob("*.xlsx")):
        if "caiq" in f.name.lower():
            return f
    return None


def is_domain_header_row(row):
    """Check if a row is a domain separator (only col 0 has value, rest empty)."""
    if not row[0]:
        return False
    non_empty = sum(1 for c in row if c)
    return non_empty == 1


def extract_aicm_controls(wb):
    """Extract controls from the AICM sheet (30 columns)."""
    ws = wb["AICM"]

    # Row 2 has group headers, Row 3 has column headers
    # Cols: 0=Domain, 1=Title, 2=ID, 3=Spec, 4=Type,
    #   5-8=Ownership (GenAI OPS/PI, Model, Orchestrated, App),
    #   9-14=Arch Relevance (Phys, Net, Compute, Storage, App, Data),
    #   15-20=Lifecycle (Prep, Dev, Eval, Deploy, Delivery, Retirement),
    #   21-29=Threat Categories
    OWNERSHIP_COLS = {5: "GenAI OPS/PI", 6: "Model", 7: "Orchestrated Services", 8: "Application"}
    ARCH_COLS = {9: "Physical", 10: "Network", 11: "Compute", 12: "Storage", 13: "Application", 14: "Data"}
    LIFECYCLE_COLS = {15: "Preparation", 16: "Development", 17: "Evaluation/Validation",
                     18: "Deployment", 19: "Delivery", 20: "Service Retirement"}
    THREAT_COLS = {21: "Model manipulation", 22: "Data poisoning",
                  23: "Sensitive data disclosure", 24: "Model theft",
                  25: "Model/Service Failure", 26: "Insecure supply chain",
                  27: "Insecure apps/plugins", 28: "Denial of Service",
                  29: "Loss of governance/compliance"}

    controls = []
    domains = OrderedDict()
    current_domain_full = ""

    for row_idx in range(4, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        # Skip empty rows
        if not any(row):
            continue

        # Domain header rows (only col 0 populated, like "Audit & Assurance - A&A")
        if is_domain_header_row(row):
            current_domain_full = str(row[0]).strip()
            continue

        # Control rows need at least ID (col 2)
        control_id = str(row[2]).strip() if row[2] else ""
        if not control_id:
            continue

        domain = str(row[0]).strip() if row[0] else ""
        title = str(row[1]).strip() if row[1] else ""
        spec = str(row[3]).strip() if row[3] else ""
        control_type = str(row[4]).strip() if row[4] else ""

        # Extract ownership
        ownership = {}
        for col_idx, label in OWNERSHIP_COLS.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val:
                ownership[label] = str(val).strip()

        # Extract architectural relevance
        arch_relevant = []
        for col_idx, label in ARCH_COLS.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val and str(val).strip().lower() == "true":
                arch_relevant.append(label)

        # Extract lifecycle relevance
        lifecycle = {}
        for col_idx, label in LIFECYCLE_COLS.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val:
                lifecycle[label] = str(val).strip()

        # Extract threat categories
        threat_cats = []
        for col_idx, label in THREAT_COLS.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val and str(val).strip().lower() == "true":
                threat_cats.append(label)

        control = {
            "id": control_id,
            "domain": domain,
            "domain_full": current_domain_full,
            "title": title,
            "specification": spec,
            "control_type": control_type,
            "ownership": ownership if ownership else None,
            "architectural_relevance": arch_relevant if arch_relevant else None,
            "lifecycle_relevance": lifecycle if lifecycle else None,
            "threat_categories": threat_cats if threat_cats else None,
            "cross_references": [],
        }

        # Remove None fields
        control = {k: v for k, v in control.items() if v is not None}

        controls.append(control)

        if domain not in domains:
            domains[domain] = []
        domains[domain].append(control)

    return controls, domains


def extract_mappings(wb, controls):
    """Extract cross-framework mappings from Scope Applicability sheet."""
    ws = wb["Scope Applicability (Mappings)"]

    # Row 2 has framework group headers: Col4=BSI, Col7=EU AI Act, Col10=ISO 42001, Col13=NIST 600-1
    # Row 3 has sub-headers: Control Mapping, Gap Level, Addendum (repeated per framework)
    FRAMEWORK_COLS = {
        "bsi_aic4": {"mapping": 4, "gap": 5, "addendum": 6},
        "eu_ai_act": {"mapping": 7, "gap": 8, "addendum": 9},
        "iso_42001": {"mapping": 10, "gap": 11, "addendum": 12},
        "nist_ai_600_1": {"mapping": 13, "gap": 14, "addendum": 15},
    }

    # Build lookup by control ID
    ctrl_by_id = {c["id"]: c for c in controls}
    mapping_count = 0

    for row_idx in range(4, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        if not any(row):
            continue

        control_id = str(row[2]).strip() if row[2] else ""
        if not control_id or control_id not in ctrl_by_id:
            continue

        ctrl = ctrl_by_id[control_id]

        for fw_name, cols in FRAMEWORK_COLS.items():
            mapping_val = row[cols["mapping"]] if cols["mapping"] < len(row) else None
            gap_val = row[cols["gap"]] if cols["gap"] < len(row) else None
            addendum_val = row[cols["addendum"]] if cols["addendum"] < len(row) else None

            if mapping_val:
                ref = {
                    "framework": fw_name,
                    "references": str(mapping_val).strip(),
                    "gap_level": str(gap_val).strip() if gap_val else "",
                }
                if addendum_val and str(addendum_val).strip() not in ("N/A", ""):
                    ref["addendum"] = str(addendum_val).strip()
                ctrl["cross_references"].append(ref)
                mapping_count += 1

    return mapping_count


def extract_implementation_guidelines(wb, controls):
    """Extract implementation guidelines and attach to controls."""
    ws = wb["Implementation Guidelines"]

    # Row 3 headers: Col4=Shared, Col5=MP, Col6=OSP, Col7=AP, Col8=AIC, Col9=CSP
    GUIDE_COLS = {
        4: "shared",
        5: "model_provider",
        6: "orchestrated_service_provider",
        7: "application_provider",
        8: "ai_customer",
        9: "cloud_service_provider",
    }

    ctrl_by_id = {c["id"]: c for c in controls}
    count = 0

    for row_idx in range(4, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        if not any(row):
            continue

        control_id = str(row[2]).strip() if row[2] else ""
        if not control_id or control_id not in ctrl_by_id:
            continue

        guidelines = {}
        for col_idx, label in GUIDE_COLS.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val:
                guidelines[label] = str(val).strip()

        if guidelines:
            ctrl_by_id[control_id]["implementation_guidelines"] = guidelines
            count += 1

    return count


def extract_auditing_guidelines(wb, controls):
    """Extract auditing guidelines and attach to controls."""
    ws = wb["Auditing Guidelines"]

    # Row 2 is headers for this sheet (different from others)
    # Col4=AP, Col5=OSP, Col6=MP, Col7=AIC, Col8=CSP
    AUDIT_COLS = {
        4: "application_provider",
        5: "orchestrated_service_provider",
        6: "model_provider",
        7: "ai_customer",
        8: "cloud_service_provider",
    }

    ctrl_by_id = {c["id"]: c for c in controls}
    count = 0

    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        if not any(row):
            continue

        control_id = str(row[2]).strip() if row[2] else ""
        if not control_id or control_id not in ctrl_by_id:
            continue

        guidelines = {}
        for col_idx, label in AUDIT_COLS.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val:
                guidelines[label] = str(val).strip()

        if guidelines:
            ctrl_by_id[control_id]["auditing_guidelines"] = guidelines
            count += 1

    return count


def extract_caiq_questions(caiq_path, controls):
    """Extract AI-CAIQ questions and attach to controls by ID prefix."""
    wb = openpyxl.load_workbook(caiq_path, data_only=True)

    # Find the main CAIQ sheet
    target = None
    for name in wb.sheetnames:
        if "caiq" in name.lower() and "intro" not in name.lower():
            target = name
            break
    if not target:
        wb.close()
        return 0

    ws = wb[target]

    # Row 2 has headers: Question ID, Question, ...
    ctrl_by_id = {c["id"]: c for c in controls}
    count = 0

    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        if not any(row):
            continue

        q_id = str(row[0]).strip() if row[0] else ""
        question = str(row[1]).strip() if row[1] else ""

        if not q_id or not question:
            continue

        # Map question to control: A&A-01.1 -> A&A-01
        parts = q_id.rsplit(".", 1)
        ctrl_id = parts[0] if len(parts) == 2 else q_id

        if ctrl_id in ctrl_by_id:
            if "caiq_questions" not in ctrl_by_id[ctrl_id]:
                ctrl_by_id[ctrl_id]["caiq_questions"] = []
            ctrl_by_id[ctrl_id]["caiq_questions"].append({
                "id": q_id,
                "question": question,
            })
            count += 1

    wb.close()
    return count


def write_json(controls, domains, output_path):
    """Write structured JSON output."""
    output = {
        "metadata": {
            "source": "https://cloudsecurityalliance.org/artifacts/ai-controls-matrix",
            "framework": "CSA AI Controls Matrix (AICM)",
            "version": "1.0.3",
            "license": "CSA License (free download, xlsx not redistributable)",
            "retrieved": "2026-04-05",
            "total_controls": len(controls),
            "total_domains": len(domains),
        },
        "domains": [
            {
                "name": domain_name,
                "control_count": len(domain_controls),
                "control_ids": [c["id"] for c in domain_controls],
            }
            for domain_name, domain_controls in domains.items()
        ],
        "controls": controls,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"Wrote {output_path} ({output_path.stat().st_size:,} bytes)")


def write_markdown(controls, domains, output_path):
    """Write human-readable markdown reference."""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("# CSA AI Controls Matrix (AICM) v1.0.3\n\n")
        f.write("Source: https://cloudsecurityalliance.org/artifacts/ai-controls-matrix\n")
        f.write("License: CSA License (free download)\n")
        f.write("Retrieved: 2026-04-05\n")
        f.write(f"Total Controls: {len(controls)}\n")
        f.write(f"Total Domains: {len(domains)}\n\n")
        f.write("---\n\n")

        # Table of contents
        f.write("# Domains\n\n")
        for i, (domain_name, domain_controls) in enumerate(domains.items(), 1):
            f.write(f"{i}. {domain_name} ({len(domain_controls)} controls)\n")
        f.write("\n---\n\n")

        # Controls by domain
        for domain_name, domain_controls in domains.items():
            f.write(f"# Domain: {domain_name}\n\n")

            for ctrl in domain_controls:
                ctrl_id = ctrl.get("id", "UNKNOWN")
                title = ctrl.get("title", "")
                spec = ctrl.get("specification", "")

                if title:
                    f.write(f"## {ctrl_id}: {title}\n\n")
                else:
                    f.write(f"## {ctrl_id}\n\n")

                f.write("**Type:** control\n")

                if ctrl.get("control_type"):
                    f.write(f"**Control Type:** {ctrl['control_type']}\n")

                if ctrl.get("architectural_relevance"):
                    f.write(f"**Architectural Relevance:** {', '.join(ctrl['architectural_relevance'])}\n")

                if ctrl.get("threat_categories"):
                    f.write(f"**Threat Categories:** {', '.join(ctrl['threat_categories'])}\n")

                # Ownership
                ownership = ctrl.get("ownership", {})
                if ownership:
                    f.write("\n**Ownership:**\n")
                    for role, val in ownership.items():
                        f.write(f"- {role}: {val}\n")

                # Lifecycle
                lifecycle = ctrl.get("lifecycle_relevance", {})
                if lifecycle:
                    f.write("\n**Lifecycle Relevance:**\n")
                    for phase, val in lifecycle.items():
                        f.write(f"- {phase}: {val}\n")

                f.write(f"\n### Specification\n\n{spec}\n\n")

                # Cross-references
                cross_refs = ctrl.get("cross_references", [])
                if cross_refs:
                    f.write("### Cross-References\n\n")
                    for ref in cross_refs:
                        fw = ref.get("framework", "")
                        refs = ref.get("references", "")
                        gap = ref.get("gap_level", "")
                        line = f"- **{fw}**: {refs}"
                        if gap:
                            line += f" (Gap: {gap})"
                        f.write(line + "\n")
                    f.write("\n")

                # CAIQ questions
                caiq = ctrl.get("caiq_questions", [])
                if caiq:
                    f.write("### CAIQ Questions\n\n")
                    for q in caiq:
                        f.write(f"- **{q['id']}**: {q['question']}\n")
                    f.write("\n")

                f.write("---\n\n")

    print(f"Wrote {output_path} ({output_path.stat().st_size:,} bytes)")


def main():
    main_xlsx = find_main_xlsx()
    caiq_xlsx = find_caiq_xlsx()

    if not main_xlsx:
        print("ERROR: No AICM .xlsx file found in data/frameworks/csa-aicm/")
        print("Download from: https://cloudsecurityalliance.org/artifacts/ai-controls-matrix")
        return

    print(f"Main AICM file: {main_xlsx.name} ({main_xlsx.stat().st_size:,} bytes)")
    if caiq_xlsx:
        print(f"AI-CAIQ file:   {caiq_xlsx.name} ({caiq_xlsx.stat().st_size:,} bytes)")

    wb = openpyxl.load_workbook(main_xlsx, data_only=True)

    # Step 1: Extract controls from AICM sheet
    print("\n--- Extracting controls from AICM sheet ---")
    controls, domains = extract_aicm_controls(wb)
    print(f"  Controls: {len(controls)}, Domains: {len(domains)}")

    # Step 2: Extract cross-framework mappings
    print("\n--- Extracting cross-framework mappings ---")
    mapping_count = extract_mappings(wb, controls)
    print(f"  Mappings added: {mapping_count}")

    # Step 3: Extract implementation guidelines
    print("\n--- Extracting implementation guidelines ---")
    impl_count = extract_implementation_guidelines(wb, controls)
    print(f"  Controls with guidelines: {impl_count}")

    # Step 4: Extract auditing guidelines
    print("\n--- Extracting auditing guidelines ---")
    audit_count = extract_auditing_guidelines(wb, controls)
    print(f"  Controls with auditing guidelines: {audit_count}")

    wb.close()

    # Step 5: Extract AI-CAIQ questions
    if caiq_xlsx:
        print("\n--- Extracting AI-CAIQ questions ---")
        caiq_count = extract_caiq_questions(caiq_xlsx, controls)
        print(f"  CAIQ questions mapped: {caiq_count}")

    # Write outputs
    print()
    write_json(controls, domains, AICM_DIR / "csa_aicm.json")
    write_markdown(controls, domains, AICM_DIR / "csa_aicm.md")

    # Summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"Controls: {len(controls)}")
    print(f"Domains:  {len(domains)}")
    print(f"\nDomain breakdown:")
    for domain_name, domain_controls in domains.items():
        print(f"  {domain_name}: {len(domain_controls)} controls")

    # Cross-reference stats
    ref_frameworks = {}
    for c in controls:
        for ref in c.get("cross_references", []):
            fw = ref.get("framework", "unknown")
            ref_frameworks[fw] = ref_frameworks.get(fw, 0) + 1
    print(f"\nCross-references by framework:")
    for fw, count in sorted(ref_frameworks.items()):
        print(f"  {fw}: {count}")
    print(f"  Total: {sum(ref_frameworks.values())}")


if __name__ == "__main__":
    main()
